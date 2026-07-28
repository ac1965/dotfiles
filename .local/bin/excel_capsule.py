#!/usr/bin/env python3
"""
excel_capsule.py  —  Excel Capsule System v2.0  (Python CLI port)

Export : .hex ファイルをチャンク分割・UUID偽装 → xlsx シートに分散保存
Restore: xlsx から指定シートを逆変換 → .hex ファイル出力

依存パッケージ:
    pip install openpyxl

使い方:
    python excel_capsule.py export  <hex_file> <xlsx_file> [--sheet SHEET]
    python excel_capsule.py restore <xlsx_file> <hex_file>  [--sheet SHEET]

VBA 版との主な変更点:
    GUI ダイアログ    → CLI 引数に置き換え
    Rnd() / Randomize → secrets.SystemRandom()（CSPRNG）
    MakeRandomUUID()  → uuid.uuid4()（標準ライブラリ）
    パディング対策    → HEX_LEN をシートのメタセル（G1）に保存して復元時にトリム
                        ※ VBA 版はこの対策なし（端数チャンクで trailing zero が残る）
"""

from __future__ import annotations

import argparse
import re
import secrets
import sys
import textwrap
import uuid
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.exit(
        "[ERROR] openpyxl が見つかりません。\n"
        "       pip install openpyxl  を実行してください。"
    )


# ---------------------------------------------------------------------------
# 定数  (VBA 版の Private Const と 1:1 対応)
# ---------------------------------------------------------------------------
CHUNK_SIZE        = 32              # チャンク長 = UUID 1個分の Hex 文字数
SCATTER_COLS      = 5               # ペイロード分散列数（B〜F 列）
DATA_START_ROW    = 2               # データ開始行（1-indexed、行1 はヘッダ）
KEY_COL           = 1               # KEY 列インデックス（A 列 = 1）
PAYLOAD_COL_START = 2               # ペイロード開始列インデックス（B 列 = 2）
TS_BASE           = 1_700_000_000   # タイムスタンプ基底（2023-11 月ごろの UNIX 時刻）
TS_STEP           = 3_600           # 行間タイムスタンプ増分（秒）
KEY_SEPARATOR     = ","

# メタデータセル（ヘッダ行の G1 = PAYLOAD_COL_START + SCATTER_COLS）
# 元 Hex 文字列長をここに記録しておき、Restore 時に trailing-zero を除去する
META_LEN_COL = PAYLOAD_COL_START + SCATTER_COLS   # 列 7 (G)
META_LEN_KEY = "hex_len"                           # G1 のラベル

# UUID 形式に見せるためのダミーヘッダ（A〜F 列）
DUMMY_HEADERS = [
    "order_id", "transaction", "reference", "session", "checksum", "trace_id",
]


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def is_hex_string(s: str) -> bool:
    """文字列が 0-9 / a-f / A-F のみで構成された非空文字列かを検証する。"""
    return bool(s) and re.fullmatch(r"[0-9a-fA-F]+", s) is not None


def hex_to_uuid(hex_str: str) -> str:
    """32 文字の Hex 文字列を UUID 形式（8-4-4-4-12）に変換する。
    最終チャンクが 32 文字未満の場合は '0' でパディングする（VBA 版と同挙動）。"""
    h = (hex_str + "0" * 32)[:32]
    return f"{h[0:8]}-{h[8:12]}-{h[12:16]}-{h[16:20]}-{h[20:32]}"


def uuid_to_hex(uuid_str: str) -> str:
    """UUID 文字列のハイフンを除去して小文字 Hex 文字列を返す。"""
    return uuid_str.replace("-", "").lower()


def make_random_uuid() -> str:
    """ダミーセル用 UUID v4 を生成する（Python 標準 uuid モジュール使用）。"""
    return str(uuid.uuid4())


# ---------------------------------------------------------------------------
# Export: .hex → xlsx
# ---------------------------------------------------------------------------

def export_capsule(hex_path: Path, xlsx_path: Path, sheet_name: str) -> None:
    """
    .hex ファイルをチャンク分割・シャッフル・UUID 偽装して xlsx に保存する。

    レイアウト:
        行1  : A〜F = ダミーヘッダ、G1 = hex_len（メタデータ）
        行2〜 : A = "ts,col_offset,chunk_idx"
                B〜F のいずれか = 実チャンクの UUID、残り = ダミー UUID v4
    """
    # ── 入力ファイル読み込み・検証 ─────────────────────────────────────────
    try:
        hex_text = hex_path.read_text(encoding="ascii").strip()
    except (OSError, UnicodeDecodeError) as e:
        sys.exit(f"[ERROR] .hex ファイルの読み込みに失敗しました: {e}")

    if not hex_text:
        sys.exit("[ERROR] ファイルが空です。")

    if not is_hex_string(hex_text):
        sys.exit(
            "[ERROR] 入力ファイルが Hex 形式ではありません。\n"
            "       secure-archive スクリプトの出力（.hex）を指定してください。"
        )

    original_len = len(hex_text)
    hex_text = hex_text.lower()   # 小文字に正規化

    # ── チャンク配列生成（0-indexed） ──────────────────────────────────────
    n_chunks = (original_len + CHUNK_SIZE - 1) // CHUNK_SIZE
    chunks = [
        hex_text[i * CHUNK_SIZE : (i + 1) * CHUNK_SIZE]
        for i in range(n_chunks)
    ]

    # ── Fisher-Yates シャッフル ─────────────────────────────────────────────
    rng = secrets.SystemRandom()
    order = list(range(n_chunks))
    rng.shuffle(order)

    # ── xlsx 取得 or 新規作成 ───────────────────────────────────────────────
    if xlsx_path.exists():
        wb = load_workbook(str(xlsx_path))
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        wb = Workbook()
        ws = wb.active           # type: ignore[assignment]
        ws.title = sheet_name

    # ── ヘッダ行書き込み（A〜F = ダミーヘッダ、G1 = hex_len メタデータ）──
    for col_idx, header in enumerate(DUMMY_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    ws.cell(row=1, column=META_LEN_COL, value=f"{META_LEN_KEY}:{original_len}")

    # ── チャンクをシャッフル順で書き込み ───────────────────────────────────
    for write_pos, chunk_zero_idx in enumerate(order):
        row_idx    = DATA_START_ROW + write_pos
        col_offset = (write_pos % SCATTER_COLS) + 1          # 1-5
        target_col = PAYLOAD_COL_START + col_offset - 1      # B=2〜F=6

        # KEY セル: "タイムスタンプ, 列オフセット, チャンクインデックス(1-based)"
        ts_val = TS_BASE + write_pos * TS_STEP + rng.randrange(60)
        ws.cell(
            row=row_idx, column=KEY_COL,
            value=f"{ts_val}{KEY_SEPARATOR}{col_offset}{KEY_SEPARATOR}{chunk_zero_idx + 1}",
        )

        # 実チャンク → UUID 形式に変換して target_col へ配置
        ws.cell(row=row_idx, column=target_col, value=hex_to_uuid(chunks[chunk_zero_idx]))

        # その他のペイロード列 → ダミー UUID v4 で埋める
        for col in range(PAYLOAD_COL_START, PAYLOAD_COL_START + SCATTER_COLS):
            if col != target_col:
                ws.cell(row=row_idx, column=col, value=make_random_uuid())

    # ── 保存 ───────────────────────────────────────────────────────────────
    try:
        wb.save(str(xlsx_path))
    except OSError as e:
        sys.exit(f"[ERROR] xlsx の保存に失敗しました: {e}")

    print(
        f"[OK] Export 完了\n"
        f"     保存先    : {xlsx_path}\n"
        f"     シート    : {sheet_name}\n"
        f"     チャンク数: {n_chunks}\n"
        f"     Hex 文字数: {original_len}"
    )


# ---------------------------------------------------------------------------
# Restore: xlsx → .hex
# ---------------------------------------------------------------------------

def restore_capsule(xlsx_path: Path, hex_path: Path, sheet_name: str) -> None:
    """
    xlsx の指定シートからチャンクを再集約し、元の .hex ファイルを復元する。

    G1 セルに hex_len メタデータがある場合はその長さにトリムする。
    メタデータがない場合（VBA 版が生成した xlsx）は trailing-zero をそのまま出力し、
    警告を表示する（VBA 版と同等の動作）。
    """
    if not xlsx_path.exists():
        sys.exit(f"[ERROR] ファイルが見つかりません: {xlsx_path}")

    try:
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        sys.exit(f"[ERROR] xlsx の読み込みに失敗しました: {e}")

    if sheet_name not in wb.sheetnames:
        wb.close()
        sys.exit(f"[ERROR] シート '{sheet_name}' が見つかりません。")

    ws = wb[sheet_name]

    # ── ヘッダ行から hex_len メタデータを読む（G1）──────────────────────────
    original_len: int | None = None
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is not None:
        meta_val = header_row[META_LEN_COL - 1] if len(header_row) >= META_LEN_COL else None
        if meta_val is not None:
            meta_str = str(meta_val)
            if meta_str.startswith(f"{META_LEN_KEY}:"):
                try:
                    original_len = int(meta_str.split(":", 1)[1])
                except ValueError:
                    pass

    # ── データ行をスキャン：chunk_map に chunk_idx → hex を収集 ─────────────
    chunk_map: dict[int, str] = {}
    n_chunks = 0

    for row_num, row in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
        start=DATA_START_ROW,
    ):
        key_val = row[KEY_COL - 1]
        if key_val is None:
            continue

        key_str = str(key_val)
        parts   = key_str.split(KEY_SEPARATOR)

        if len(parts) < 3:
            wb.close()
            sys.exit(f"[ERROR] KEY 形式が不正です。(行: {row_num}, 値: {key_str!r})")

        try:
            col_offset = int(parts[1])
            chunk_idx  = int(parts[2])
        except ValueError:
            wb.close()
            sys.exit(f"[ERROR] KEY のパースに失敗しました。(行: {row_num}, 値: {key_str!r})")

        if not (1 <= col_offset <= SCATTER_COLS):
            wb.close()
            sys.exit(
                f"[ERROR] 列オフセットが範囲外です。"
                f"(行: {row_num}, col_offset: {col_offset}, 期待値: 1-{SCATTER_COLS})"
            )

        if chunk_idx < 1:
            wb.close()
            sys.exit(f"[ERROR] チャンクインデックスが不正です。(行: {row_num}, 値: {chunk_idx})")

        payload_col_zero = PAYLOAD_COL_START - 1 + col_offset - 1
        uuid_val = row[payload_col_zero]
        if uuid_val is None:
            wb.close()
            sys.exit(f"[ERROR] ペイロードセルが空です。(行: {row_num})")

        chunk_map[chunk_idx] = uuid_to_hex(str(uuid_val))
        if chunk_idx > n_chunks:
            n_chunks = chunk_idx

    wb.close()

    if n_chunks <= 0:
        sys.exit("[ERROR] 有効なチャンクインデックスが見つかりませんでした。")

    # ── 欠損チャンク確認 ─────────────────────────────────────────────────────
    missing = [i for i in range(1, n_chunks + 1) if i not in chunk_map]
    if missing:
        sys.exit(f"[ERROR] 以下のチャンクが見つかりません: {missing}")

    # ── chunk_idx 昇順で連結 ─────────────────────────────────────────────────
    hex_text = "".join(chunk_map[i] for i in range(1, n_chunks + 1))

    # ── trailing-zero トリム（端数チャンクのパディング除去） ─────────────────
    if original_len is not None:
        # このファイルで生成された xlsx → 正確にトリムできる
        hex_text = hex_text[:original_len]
    elif len(hex_text) % CHUNK_SIZE != 0:
        # VBA 版生成の xlsx など、メタデータなし → 警告のみ
        print(
            "[WARN] G1 に hex_len メタデータがありません。\n"
            "       端数チャンクが 0 でパディングされた状態で出力します（VBA 版と同等）。",
            file=sys.stderr,
        )

    # ── .hex ファイルに書き込み（改行なし、バイナリモード）──────────────────
    try:
        hex_path.parent.mkdir(parents=True, exist_ok=True)
        hex_path.write_bytes(hex_text.encode("ascii"))
    except OSError as e:
        sys.exit(f"[ERROR] .hex ファイルの書き込みに失敗しました: {e}")

    print(
        f"[OK] Restore 完了\n"
        f"     保存先    : {hex_path}\n"
        f"     チャンク数: {n_chunks}\n"
        f"     Hex 文字数: {len(hex_text)}"
    )


# ---------------------------------------------------------------------------
# インタラクティブウィザード（引数なし起動時）
# ---------------------------------------------------------------------------

def _prompt(label: str, default: str = "") -> str:
    """プロンプトを表示して入力を受け取る。空 Enter でデフォルト値を返す。"""
    hint = f"  [{default}]" if default else ""
    try:
        val = input(f"  {label}{hint}: ").strip()
    except (EOFError, KeyboardInterrupt):
        print("\n中断しました。")
        sys.exit(0)
    return val if val else default


def _prompt_path(label: str, default: str = "") -> str:
    """パスを入力させる。シェルのチルダ展開を行う。"""
    import os
    raw = _prompt(label, default)
    return os.path.expanduser(raw) if raw else raw


def _choose_command() -> str:
    """e / r のどちらかを選ばせる。"""
    print("  操作を選んでください:")
    print("    e) Export  — .hex → xlsx に格納")
    print("    r) Restore — xlsx → .hex に復元")
    while True:
        try:
            ch = input("  [e/r]: ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            print("\n中断しました。")
            sys.exit(0)
        if ch in ("e", "export"):
            return "export"
        if ch in ("r", "restore"):
            return "restore"
        print("  'e' または 'r' を入力してください。")


def run_interactive() -> None:
    """
    引数なし起動時のウィザードモード。
    VBA 版の InputBox / FileDialog 相当をターミナル上で再現する。
    """
    print("=" * 56)
    print("  Excel Capsule System v2.0  —  Interactive Mode")
    print("=" * 56)

    command = _choose_command()
    print()

    if command == "export":
        hex_file  = _prompt_path(".hex ファイルのパス（入力）")
        xlsx_file = _prompt_path(".xlsx ファイルのパス（出力）", "capsule.xlsx")
        sheet     = _prompt("シート名", "transactions")
        print()
        export_capsule(Path(hex_file), Path(xlsx_file), sheet)

    else:  # restore
        xlsx_file = _prompt_path(".xlsx ファイルのパス（入力）")
        hex_file  = _prompt_path(".hex ファイルのパス（出力）", "output.hex")
        sheet     = _prompt("シート名", "transactions")
        print()
        restore_capsule(Path(xlsx_file), Path(hex_file), sheet)


# ---------------------------------------------------------------------------
# CLI エントリポイント
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="excel_capsule",
        description="Excel Capsule System v2.0 — .hex ↔ xlsx 変換 CLI",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            引数なしで起動するとインタラクティブウィザードが起動します。

            使用例（非対話）:
              python excel_capsule.py export  secret.hex capsule.xlsx --sheet transactions
              python excel_capsule.py restore capsule.xlsx restored.hex --sheet transactions
        """),
    )

    sub = parser.add_subparsers(dest="command")   # required=False にして引数なしを許容

    # ── export ─────────────────────────────────────────────────────────────
    exp = sub.add_parser("export",  help=".hex → xlsx（チャンク分散・UUID 偽装）")
    exp.add_argument("hex_file",  help="入力 .hex ファイルのパス")
    exp.add_argument("xlsx_file", help="出力 .xlsx ファイルのパス")
    exp.add_argument("--sheet", default="transactions", metavar="SHEET",
                     help="書き込み先シート名  [デフォルト: transactions]")

    # ── restore ────────────────────────────────────────────────────────────
    res = sub.add_parser("restore", help="xlsx → .hex（チャンク再集約・Hex 復元）")
    res.add_argument("xlsx_file", help="読み込む .xlsx ファイルのパス")
    res.add_argument("hex_file",  help="出力 .hex ファイルのパス")
    res.add_argument("--sheet", default="transactions", metavar="SHEET",
                     help="読み込むシート名  [デフォルト: transactions]")

    return parser


def main() -> None:
    parser = build_parser()
    args   = parser.parse_args()

    # 引数なし → インタラクティブウィザード
    if args.command is None:
        run_interactive()
        return

    if args.command == "export":
        export_capsule(
            hex_path   = Path(args.hex_file),
            xlsx_path  = Path(args.xlsx_file),
            sheet_name = args.sheet,
        )
    else:
        restore_capsule(
            xlsx_path  = Path(args.xlsx_file),
            hex_path   = Path(args.hex_file),
            sheet_name = args.sheet,
        )


if __name__ == "__main__":
    main()
